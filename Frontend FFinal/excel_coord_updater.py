"""
Excel Coordinate Updater for School Transport Management
This script generates and updates Excel files with vehicle coordinates in real-time
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import json
import time
import os
from datetime import datetime
import random

class ExcelCoordUpdater:
    def __init__(self, filename='vehicle_coordinates.xlsx'):
        self.filename = filename
        self.workbook = None
        self.worksheet = None
        self.init_excel()
    
    def init_excel(self):
        """Initialize or load Excel file"""
        if os.path.exists(self.filename):
            # Load existing workbook
            self.workbook = openpyxl.load_workbook(self.filename)
            if 'Vehicle Data' in self.workbook.sheetnames:
                self.worksheet = self.workbook['Vehicle Data']
            else:
                self.worksheet = self.workbook.create_sheet('Vehicle Data')
                self.create_headers()
        else:
            # Create new workbook
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = 'Vehicle Data'
            self.create_headers()
            self.workbook.save(self.filename)
            print(f"Created new Excel file: {self.filename}")
    
    def create_headers(self):
        """Create header row with styling"""
        headers = [
            'Vehicle ID', 'Vehicle Name', 'Driver', 'Route',
            'Latitude', 'Longitude', 'Speed (km/h)', 'Max Speed (km/h)',
            'Avg Speed (km/h)', 'Status', 'Engine Status',
            'Timestamp', 'Date', 'Time'
        ]
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        column_widths = {
            'A': 12, 'B': 18, 'C': 15, 'D': 12,
            'E': 12, 'F': 12, 'G': 12, 'H': 14,
            'I': 14, 'J': 12, 'K': 14, 'L': 20,
            'M': 12, 'N': 10
        }
        
        for col, width in column_widths.items():
            self.worksheet.column_dimensions[col].width = width
        
        # Freeze header row
        self.worksheet.freeze_panes = 'A2'
    
    def update_vehicle_data(self, vehicle_data):
        """
        Update Excel file with vehicle coordinates and data
        
        Args:
            vehicle_data: List of dictionaries containing vehicle information
        """
        # Find the next empty row
        next_row = self.worksheet.max_row + 1
        
        # Status color coding
        status_colors = {
            'moving': '90EE90',      # Light green
            'stopped': 'D3D3D3',     # Light gray
            'overspeed': 'FF6B6B'    # Light red
        }
        
        for vehicle in vehicle_data:
            now = datetime.now()
            
            row_data = [
                vehicle.get('id', ''),
                vehicle.get('name', ''),
                vehicle.get('driver', ''),
                vehicle.get('route', ''),
                round(vehicle.get('lat', 0), 6),
                round(vehicle.get('lng', 0), 6),
                round(vehicle.get('speed', 0), 2),
                round(vehicle.get('maxSpeed', 0), 2),
                round(vehicle.get('avgSpeed', 0), 2),
                vehicle.get('status', '').upper(),
                vehicle.get('engineStatus', '').upper(),
                now.isoformat(),
                now.strftime('%Y-%m-%d'),
                now.strftime('%H:%M:%S')
            ]
            
            # Write data to row
            for col_num, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=next_row, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Color code status column
                if col_num == 10:  # Status column
                    status = vehicle.get('status', '').lower()
                    if status in status_colors:
                        cell.fill = PatternFill(
                            start_color=status_colors[status],
                            end_color=status_colors[status],
                            fill_type="solid"
                        )
            
            next_row += 1
        
        # Save workbook
        self.workbook.save(self.filename)
        print(f"Updated Excel file: {self.filename} at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    def generate_sample_data(self, num_vehicles=3):
        """Generate sample vehicle data for testing"""
        vehicles = []
        
        # Base coordinates (Visakhapatnam area)
        base_lat = 17.6868
        base_lng = 83.2185
        
        vehicle_info = [
            {'id': 'BUS-001', 'name': 'School Bus 1', 'driver': 'John Doe', 'route': 'Route A'},
            {'id': 'BUS-002', 'name': 'School Bus 2', 'driver': 'Jane Smith', 'route': 'Route B'},
            {'id': 'BUS-003', 'name': 'School Bus 3', 'driver': 'Mike Johnson', 'route': 'Route C'}
        ]
        
        for i, info in enumerate(vehicle_info[:num_vehicles]):
            # Simulate movement with slight variations
            lat = base_lat + (random.random() - 0.5) * 0.01
            lng = base_lng + (random.random() - 0.5) * 0.01
            
            # Random speed between 0-80 km/h
            speed = random.uniform(0, 80)
            
            # Determine status based on speed
            if speed > 60:
                status = 'overspeed'
            elif speed > 0:
                status = 'moving'
            else:
                status = 'stopped'
            
            vehicle = {
                'id': info['id'],
                'name': info['name'],
                'driver': info['driver'],
                'route': info['route'],
                'lat': lat,
                'lng': lng,
                'speed': round(speed, 2),
                'maxSpeed': round(random.uniform(speed, 85), 2),
                'avgSpeed': round(speed * 0.8, 2),
                'status': status,
                'engineStatus': 'on' if speed > 0 else 'off'
            }
            
            vehicles.append(vehicle)
        
        return vehicles
    
    def run_continuous_updates(self, interval=5):
        """
        Continuously update Excel file with new coordinates
        
        Args:
            interval: Update interval in seconds
        """
        print(f"Starting continuous updates every {interval} seconds...")
        print("Press Ctrl+C to stop")
        
        try:
            while True:
                # Generate or receive vehicle data
                # In production, this would come from your vehicle tracking system
                vehicle_data = self.generate_sample_data()
                
                # Update Excel file
                self.update_vehicle_data(vehicle_data)
                
                # Wait for next update
                time.sleep(interval)
        
        except KeyboardInterrupt:
            print("\nStopping updates...")
            self.workbook.save(self.filename)
            print("Excel file saved successfully")
    
    def export_summary_report(self, output_filename='vehicle_summary.xlsx'):
        """Export a summary report of vehicle data"""
        if self.worksheet.max_row <= 1:
            print("No data to export")
            return
        
        summary_wb = Workbook()
        summary_ws = summary_wb.active
        summary_ws.title = 'Summary'
        
        # Get unique vehicles
        vehicles = {}
        for row in range(2, self.worksheet.max_row + 1):
            vehicle_id = self.worksheet.cell(row=row, column=1).value
            if vehicle_id not in vehicles:
                vehicles[vehicle_id] = {
                    'id': vehicle_id,
                    'name': self.worksheet.cell(row=row, column=2).value,
                    'max_speed': 0,
                    'total_updates': 0,
                    'last_update': None
                }
            
            vehicles[vehicle_id]['total_updates'] += 1
            speed = self.worksheet.cell(row=row, column=7).value or 0
            if speed > vehicles[vehicle_id]['max_speed']:
                vehicles[vehicle_id]['max_speed'] = speed
            
            timestamp = self.worksheet.cell(row=row, column=12).value
            if not vehicles[vehicle_id]['last_update'] or timestamp > vehicles[vehicle_id]['last_update']:
                vehicles[vehicle_id]['last_update'] = timestamp
        
        # Write summary
        summary_headers = ['Vehicle ID', 'Vehicle Name', 'Max Speed (km/h)', 'Total Updates', 'Last Update']
        for col_num, header in enumerate(summary_headers, 1):
            cell = summary_ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        row_num = 2
        for vehicle in vehicles.values():
            summary_ws.cell(row=row_num, column=1).value = vehicle['id']
            summary_ws.cell(row=row_num, column=2).value = vehicle['name']
            summary_ws.cell(row=row_num, column=3).value = round(vehicle['max_speed'], 2)
            summary_ws.cell(row=row_num, column=4).value = vehicle['total_updates']
            summary_ws.cell(row=row_num, column=5).value = vehicle['last_update']
            row_num += 1
        
        summary_wb.save(output_filename)
        print(f"Summary report exported to: {output_filename}")


def main():
    """Main function to run the Excel coordinate updater"""
    updater = ExcelCoordUpdater('vehicle_coordinates.xlsx')
    
    # Option 1: Run continuous updates (for real-time tracking)
    # updater.run_continuous_updates(interval=5)
    
    # Option 2: Generate initial sample data
    print("Generating initial sample data...")
    sample_data = updater.generate_sample_data(num_vehicles=3)
    updater.update_vehicle_data(sample_data)
    
    # Option 3: Export summary report
    # updater.export_summary_report()
    
    print("Excel file ready for use with the web dashboard!")


if __name__ == '__main__':
    main()






