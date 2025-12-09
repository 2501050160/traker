"""
Create Excel Template for Manual Coordinate Updates
This creates a template file that users can manually update with coordinates
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_path_template():
    """Create Excel template for manual path/coordinate updates"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Vehicle Paths'
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Vehicle ID', 'Route Name', 'Point Order', 'Latitude', 'Longitude',
        'Speed (km/h)', 'Stop Duration (sec)', 'Description'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Set column widths
    column_widths = {
        'A': 15, 'B': 18, 'C': 12, 'D': 15,
        'E': 15, 'F': 15, 'G': 18, 'H': 30
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add sample data rows
    sample_data = [
        ['BUS-001', 'Route A', 1, 17.6868, 83.2185, 0, 0, 'Starting Point'],
        ['BUS-001', 'Route A', 2, 17.6900, 83.2200, 40, 0, 'Checkpoint 1'],
        ['BUS-001', 'Route A', 3, 17.6950, 83.2250, 45, 0, 'Checkpoint 2'],
        ['BUS-001', 'Route A', 4, 17.7000, 83.2300, 0, 60, 'Destination - School'],
        ['BUS-002', 'Route B', 1, 17.6800, 83.2100, 0, 0, 'Starting Point'],
        ['BUS-002', 'Route B', 2, 17.6850, 83.2150, 50, 0, 'Checkpoint 1'],
        ['BUS-002', 'Route B', 3, 17.6900, 83.2200, 0, 120, 'Destination - School'],
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    # Instructions sheet
    ws2 = wb.create_sheet('Instructions')
    instructions = [
        ['INSTRUCTIONS FOR MANUAL UPDATES'],
        [''],
        ['1. Update coordinates in the "Vehicle Paths" sheet'],
        ['2. Each row represents a point in the vehicle route'],
        ['3. Point Order: Sequence number for the route (1, 2, 3, ...)'],
        ['4. Latitude/Longitude: GPS coordinates for each point'],
        ['5. Speed (km/h): Speed at this point (0 = stop)'],
        ['6. Stop Duration (sec): How long to wait at this point'],
        ['7. Description: Optional note about the location'],
        [''],
        ['IMPORTANT:'],
        ['- Save the file after making changes'],
        ['- Click "Refresh" button in the dashboard to load new coordinates'],
        ['- The system will automatically create route from initial to final point'],
        ['- Make sure Point Order is sequential for each vehicle'],
    ]
    
    for row_num, instruction in enumerate(instructions, 1):
        cell = ws2.cell(row=row_num, column=1)
        cell.value = instruction[0]
        if row_num == 1:
            cell.font = Font(bold=True, size=14)
        elif 'IMPORTANT' in str(instruction[0]):
            cell.font = Font(bold=True, color="FF0000")
    
    ws2.column_dimensions['A'].width = 80
    
    # Save file
    filename = 'vehicle_path_template.xlsx'
    wb.save(filename)
    print(f"Excel template created: {filename}")
    print("You can now manually update coordinates in this file!")

if __name__ == '__main__':
    create_path_template()





